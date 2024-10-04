from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os

data = [
    {
        'name': '王小名',
        'age': 36,
        'weight': 80
    },
    {
        'name': '陳小華',
        'age': 25,
        'weight': 90
    },
    {
        'name': '張小美',
        'age': 18,
        'weight': 48
    }
]

if os.path.exists('BMI.xlsx'):
    wb = load_workbook('BMI.xlsx')
else:
    wb = Workbook()

ws = wb.active

title = ['姓名', '年齡', '體重']
ws.append(title)

for p in data:
    ws.append(list(p.values()))

for col in range(2, 4):  
    char = get_column_letter(col)
    # 添加EXCEL公式
    ws[char + '7'] = f'=AVERAGE({char + "2"}: {char + "6"})'

for col in range(1, 5):
    char = get_column_letter(col)
    ws[char+'1'].font = Font(bold=True, color='000000FF')

wb.save('BMI.xlsx')
