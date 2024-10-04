# 載入模組
# import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# 開啟 EXCEL
# wb = load_workbook('excel.xlsx')
# 指定開啟工作表
# ws = wb.active
# print(ws)

# 建議寫法 避免檔案已存在 或是 未安全儲存/關閉檔案 造成錯誤
if os.path.exists('excel.xlsx'):
    wb = load_workbook('excel.xlsx')
else:
    wb = Workbook() # 創建excel 檔案

ws = wb.active # 指定使用當前的工作表
ws.title = 'sheet01' # 修改工作表名稱

print(wb.sheetnames) # 顯示所有工作表
# ws = wb['工作表名稱'] # 指定使用工作表
# wb.create_sheet('test') # 創建工作表
# print(ws['B3'].value) # 顯示指定欄位的值

ws['A5'].value = 'test'
# ws.append([123,456,789,0]) # 新增一橫排的值

# 讀取一範圍資料
for row in range(1,6):
    for col in range(1,6):
        char = get_column_letter(col)
        ws[char+str(row)].value = char + str(row)
        print(ws[char+str(row)].value) # 回傳範圍內的值


ws.merge_cells('A10:E10') # 合併儲存格
# ws.unmerge_cells('A10:E10') # 解除合併

# 插入 row / col
# ws.insert_rows(3)
# ws.insert_cols(3)

# 刪除 row / col
# ws.delete_rows(3)
# ws.delete_cols(3)

# 移動範圍資料
# ws.move_range('A3:E5', rows=2, cols=2)


wb.save('excel.xlsx') # 存檔並關閉檔案 請確認在save()執行時 檔案是關閉的

